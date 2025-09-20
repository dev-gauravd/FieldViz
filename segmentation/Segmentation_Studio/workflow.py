import cv2
import os
from ultralytics import YOLO
import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import io
import numpy as np
import pytesseract
import sys
import json
import shutil
from pathlib import Path

# Configure Tesseract path
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def process_selected_segments(selected_image_path, output_dir, models_dir):
    """
    Process a single selected segment image through column and row segmentation,
    then generate Excel output.
    
    Args:
        selected_image_path (str): Path to the selected segment image
        output_dir (str): Base output directory for all results
        models_dir (str): Directory containing the model files
    
    Returns:
        dict: Status and paths of generated files
    """
    
    try:
        # Create timestamped output directories
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_output = os.path.join(output_dir, f"export_{timestamp}")
        
        column_output_dir = os.path.join(base_output, "column_segment")
        row_output_dir = os.path.join(base_output, "row_segments")
        excel_output_dir = os.path.join(base_output, "Excel")
        
        # Create directories
        os.makedirs(column_output_dir, exist_ok=True)
        os.makedirs(row_output_dir, exist_ok=True)
        os.makedirs(excel_output_dir, exist_ok=True)
        
        # Model paths
        column_model_path = os.path.join(models_dir, "column_detect.pt")
        row_model_path = os.path.join(models_dir, "row_detect.pt")
        
        # Verify model files exist
        if not os.path.exists(column_model_path):
            raise FileNotFoundError(f"Column detection model not found: {column_model_path}")
        if not os.path.exists(row_model_path):
            raise FileNotFoundError(f"Row detection model not found: {row_model_path}")
        
        print(f"🔄 Processing selected image: {selected_image_path}")
        print(f"📁 Output directory: {base_output}")
        
        # ======================= STEP 1: COLUMN SEGMENTATION =======================
        print("🔄 Starting Column Segmentation...")
        
        # Define classes and colors for columns
        class_names = [f"c_{i}" for i in range(1, 34)]  # c_1 … c_33
        random.seed(42)
        class_colors = {
            cls: (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
            for cls in class_names
        }
        
        # Load column detection model
        column_model = YOLO(column_model_path)
        
        # Read input image
        img = cv2.imread(selected_image_path)
        if img is None:
            raise ValueError(f"Could not read image: {selected_image_path}")
        
        h, w = img.shape[:2]
        
        # Run column detection
        results = column_model(
            selected_image_path,
            conf=0.15,
            iou=0.3,
            imgsz=max(640, max(w, h)),
            max_det=100,
            augment=True,
            agnostic_nms=True,
            verbose=True
        )
        
        print(f"📊 Image dimensions: {w}x{h}")
        print(f"🔍 Total column detections found: {len(results[0].boxes) if results[0].boxes is not None else 0}")
        
        # Process column results
        annotated_img = img.copy()
        detection_count = 0
        
        for result in results:
            if result.boxes is None:
                print("⚠️ No column detections found!")
                continue
            
            boxes = result.boxes
            print(f"📦 Processing {len(boxes)} column detections...")
            
            # Sort boxes by confidence
            if len(boxes) > 0:
                confidences = boxes.conf.cpu().numpy()
                sorted_indices = np.argsort(confidences)[::-1]
                
                for idx in sorted_indices:
                    box = boxes[idx]
                    cls_id = int(box.cls[0])
                    confidence = float(box.conf[0])
                    
                    if confidence < 0.15:
                        continue
                    
                    label = result.names[cls_id] if cls_id in result.names else f"c_{cls_id+1}"
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    
                    box_w, box_h = x2 - x1, y2 - y1
                    img_area = img.shape[0] * img.shape[1]
                    box_area = box_w * box_h
                    
                    # Filter out too small or too large boxes
                    if box_area < 0.001 * img_area or box_area > 0.8 * img_area:
                        continue
                    
                    aspect_ratio = box_w / max(box_h, 1)
                    if aspect_ratio > 10 or aspect_ratio < 0.02:
                        continue
                    
                    detection_count += 1
                    
                    # Save segmented column
                    crop = img[y1:y2, x1:x2]
                    save_path = os.path.join(column_output_dir, f"{label}_conf{confidence:.2f}.png")
                    cv2.imwrite(save_path, crop)
                    
                    # Draw bounding box
                    color = class_colors.get(label, (0, 255, 255))
                    cv2.rectangle(annotated_img, (x1, y1), (x2, y2), color, 2)
                    display_label = f"{label} ({confidence:.2f})"
                    cv2.putText(
                        annotated_img, display_label, (x1, y1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2
                    )
        
        print(f"✅ Valid column detections processed: {detection_count}")
        
        # Save annotated column image
        annotated_path = os.path.join(column_output_dir, "annotated_columns.png")
        cv2.imwrite(annotated_path, annotated_img)
        
        # Rename column crops sequentially
        rename_columns_sequentially(column_output_dir, prefix="c_", total_columns=33)
        
        # ======================= STEP 2: ROW SEGMENTATION =======================
        print("🔄 Starting Row Segmentation...")
        
        # Define classes and colors for rows
        row_class_names = [f"r_{i}" for i in range(1, 25)]  # r_1 ... r_24
        random.seed(123)
        row_class_colors = {
            cls: (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
            for cls in row_class_names
        }
        
        # Load row detection model
        row_model = YOLO(row_model_path)
        
        # Process each column image for row detection
        for file_name in os.listdir(column_output_dir):
            if not (file_name.lower().endswith((".png", ".jpg", ".jpeg")) and file_name.startswith("c_")):
                continue
            
            image_path = os.path.join(column_output_dir, file_name)
            img = cv2.imread(image_path)
            if img is None:
                print(f"⚠️ Could not read {file_name}, skipping...")
                continue
            
            h, w = img.shape[:2]
            
            # Run row detection
            results = row_model(
                image_path,
                conf=0.15,
                iou=0.3,
                imgsz=max(640, max(w, h)),
                max_det=200,
                augment=True,
                agnostic_nms=True,
                verbose=False
            )
            
            print(f"\n📊 Processing {file_name} for rows ({w}x{h})")
            print(f"🔍 Total row detections: {len(results[0].boxes) if results[0].boxes is not None else 0}")
            
            annotated_img = img.copy()
            row_detection_count = 0
            row_detections = []
            
            for result in results:
                if result.boxes is None:
                    continue
                
                boxes = result.boxes
                confidences = boxes.conf.cpu().numpy()
                sorted_indices = np.argsort(confidences)[::-1]
                
                for idx in sorted_indices:
                    box = boxes[idx]
                    cls_id = int(box.cls[0])
                    conf = float(box.conf[0])
                    
                    if conf < 0.15:
                        continue
                    
                    label = result.names[cls_id] if cls_id in result.names else f"r_{cls_id+1}"
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    
                    # Row-specific filtering
                    box_w, box_h = x2 - x1, y2 - y1
                    img_area = w * h
                    box_area = box_w * box_h
                    
                    if box_area < 0.001 * img_area or box_area > 0.5 * img_area:
                        continue
                    
                    aspect_ratio = box_w / max(box_h, 1)
                    if aspect_ratio < 1.5:  # rows should be wide
                        continue
                    
                    row_detection_count += 1
                    row_detections.append((y1, x1, x2, y2, label, conf))
            
            # Sort rows top-to-bottom
            row_detections = sorted(row_detections, key=lambda x: x[0])
            
            # Save row crops for this column
            base_name = os.path.splitext(file_name)[0]
            save_dir = os.path.join(row_output_dir, base_name)
            os.makedirs(save_dir, exist_ok=True)
            
            for idx, (y1, x1, x2, y2, label, conf) in enumerate(row_detections, start=1):
                crop = img[y1:y2, x1:x2]
                save_path = os.path.join(save_dir, f"{label}_conf{conf:.2f}.png")
                cv2.imwrite(save_path, crop)
                
                # Annotate
                color = row_class_colors.get(label, (0, 255, 255))
                cv2.rectangle(annotated_img, (x1, y1), (x2, y2), color, 2)
                cv2.putText(annotated_img, f"{label} ({conf:.2f})", (x1, y1 - 10),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
            
            # Save annotated row image
            annotated_row_path = os.path.join(row_output_dir, f"{base_name}_rows_annotated.png")
            cv2.imwrite(annotated_row_path, annotated_img)
            
            print(f"✅ {row_detection_count} valid rows saved for {file_name}")
            
            # Rename rows sequentially
            rename_rows_sequentially(save_dir, prefix="r_", total_rows=24)
        
        # ======================= STEP 3: EXCEL GENERATION =======================
        print("🔄 Starting Excel Generation...")
        
        excel_file_path = generate_excel_from_segments(row_output_dir, excel_output_dir)
        
        return {
            "status": "success",
            "message": "Excel export completed successfully",
            "excel_path": excel_file_path,
            "output_dir": base_output,
            "column_segments": detection_count,
            "timestamp": timestamp
        }
        
    except Exception as e:
        print(f"❌ Error in process_selected_segments: {str(e)}")
        return {
            "status": "error",
            "message": f"Error processing segments: {str(e)}",
            "excel_path": None
        }


def rename_columns_sequentially(folder, prefix="c_", total_columns=33):
    """
    Renames all files starting with prefix (e.g., c_) in sequential order.
    """
    files = [f for f in os.listdir(folder) if f.startswith(prefix)]
    files.sort()
    
    if not files:
        print("⚠️ No column files found to rename.")
        return
    
    for idx, old_name in enumerate(files, start=1):
        if idx > total_columns:
            break
        ext = os.path.splitext(old_name)[1]
        new_name = f"{prefix}{idx}{ext}"
        old_path = os.path.join(folder, old_name)
        new_path = os.path.join(folder, new_name)
        
        if os.path.exists(new_path):
            os.remove(new_path)
        os.rename(old_path, new_path)
        print(f"🔄 Renamed {old_name} ➝ {new_name}")
    
    print(f"✅ Renamed {min(len(files), total_columns)} columns sequentially")


def rename_rows_sequentially(folder, prefix="r_", total_rows=24):
    """
    Renames all files starting with prefix (e.g., r_) in sequential order.
    """
    files = [f for f in os.listdir(folder) if f.startswith(prefix)]
    files.sort()
    
    if not files:
        print(f"⚠️ No row files found in {folder}")
        return
    
    for idx, old_name in enumerate(files, start=1):
        if idx > total_rows:
            break
        ext = os.path.splitext(old_name)[1]
        new_name = f"{prefix}{idx}{ext}"
        old_path = os.path.join(folder, old_name)
        new_path = os.path.join(folder, new_name)
        
        if os.path.exists(new_path):
            os.remove(new_path)
        os.rename(old_path, new_path)
    
    print(f"   ✅ Renamed {min(len(files), total_rows)} rows sequentially")


def ocr_digits_only(image_path):
    """
    Perform OCR restricted to digits and '.' sign only.
    """
    try:
        custom_config = r'--oem 3 --psm 7 -c tessedit_char_whitelist=0123456789.'
        text = pytesseract.image_to_string(Image.open(image_path), config=custom_config)
        text = text.strip()
        return text
    except Exception as e:
        print(f"⚠️ OCR failed on {image_path}: {e}")
        return ""


def generate_excel_from_segments(row_segments_dir, excel_output_dir):
    """
    Generate Excel file from row segments with OCR and image insertion.
    """
    # Create new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Table Data"
    
    # Get all column folders from row_segments directory
    column_folders = []
    for item in os.listdir(row_segments_dir):
        item_path = os.path.join(row_segments_dir, item)
        if os.path.isdir(item_path) and item.startswith('c_'):
            column_folders.append(item)
    
    # Sort column folders numerically
    column_folders.sort(key=lambda x: int(x.split('_')[1]) if len(x.split('_')) > 1 and x.split('_')[1].isdigit() else 0)
    
    max_cols = len(column_folders)
    row_height = 25
    col_width = 20
    temp_images = []
    
    print(f"📊 Found {max_cols} columns: {column_folders}")
    
    # Process each column
    for col_idx, column_name in enumerate(column_folders, 1):
        column_path = os.path.join(row_segments_dir, column_name)
        
        # Set column width
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_width
        
        # Add column header
        ws.cell(row=1, column=col_idx, value=column_name.upper())
        
        print(f"🔄 Processing column: {column_name}")
        
        # Get all row images in this column folder
        if os.path.exists(column_path):
            row_files = [f for f in os.listdir(column_path) if f.endswith(('.png', '.jpg'))]
            
            # Process each row image
            for row_file in row_files:
                row_name = os.path.splitext(row_file)[0]
                if row_name.startswith('r_'):
                    try:
                        # Extract row number
                        row_parts = row_name.split('_')
                        if len(row_parts) > 1 and row_parts[1].isdigit():
                            row_num = int(row_parts[1])
                            excel_row = row_num + 1  # +1 because row 1 is header
                            
                            # Set row height
                            ws.row_dimensions[excel_row].height = row_height
                            
                            image_path = os.path.join(column_path, row_file)
                            
                            # Perform OCR
                            ocr_text = ocr_digits_only(image_path)
                            
                            if ocr_text:
                                # Write OCR result into Excel cell
                                ws.cell(row=excel_row, column=col_idx, value=ocr_text)
                                print(f"  🔢 OCR extracted '{ocr_text}' from {row_name}")
                            else:
                                # Fallback: insert image into Excel
                                try:
                                    pil_img = Image.open(image_path)
                                    
                                    # Resize image to fit in cell
                                    max_width = 150
                                    max_height = 80
                                    pil_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                                    
                                    # Save resized image temporarily
                                    temp_img_path = os.path.join(excel_output_dir, f"temp_{column_name}_{row_name}.png")
                                    pil_img.save(temp_img_path)
                                    temp_images.append(temp_img_path)
                                    
                                    img = OpenpyxlImage(temp_img_path)
                                    cell_ref = ws.cell(row=excel_row, column=col_idx).coordinate
                                    img.anchor = cell_ref
                                    ws.add_image(img)
                                    
                                    print(f"  🖼️ Inserted image for {row_name} (no valid OCR)")
                                
                                except Exception as e:
                                    print(f"  ⚠️ Error processing image {row_file}: {e}")
                                    ws.cell(row=excel_row, column=col_idx, value=f"Image: {row_name}")
                        
                    except ValueError as e:
                        print(f"  ⚠️ Invalid row format: {row_name} - {e}")
        else:
            print(f"  ⚠️ Column folder not found: {column_path}")
    
    # Save Excel file
    excel_file_path = os.path.join(excel_output_dir, "table_data.xlsx")
    wb.save(excel_file_path)
    
    # Clean up temporary images
    print("🧹 Cleaning up temporary files...")
    for temp_img_path in temp_images:
        try:
            if os.path.exists(temp_img_path):
                os.remove(temp_img_path)
        except Exception as e:
            print(f"  ⚠️ Could not remove temp file {temp_img_path}: {e}")
    
    print(f"✅ Excel file saved: {excel_file_path}")
    return excel_file_path


# Main execution function
def main():
    """
    Main function to handle command line arguments and process segments.
    Expected arguments: selected_image_path, output_dir, models_dir
    """
    if len(sys.argv) != 4:
        print("Usage: python workflow.py <selected_image_path> <output_dir> <models_dir>")
        sys.exit(1)
    
    selected_image_path = sys.argv[1]
    output_dir = sys.argv[2]
    models_dir = sys.argv[3]
    
    # Verify input image exists
    if not os.path.exists(selected_image_path):
        result = {
            "status": "error",
            "message": f"Selected image not found: {selected_image_path}"
        }
        print(json.dumps(result))
        sys.exit(1)
    
    # Process the selected segments
    result = process_selected_segments(selected_image_path, output_dir, models_dir)
    
    # Output result as JSON for easy parsing by Node.js
    print(json.dumps(result))
    
    if result["status"] == "success":
        print("\n" + "="*60)
        print("📋 EXCEL EXPORT COMPLETED SUCCESSFULLY!")
        print("="*60)
        print(f"📄 Excel file: {result['excel_path']}")
        print(f"📁 Output directory: {result['output_dir']}")
        print(f"🔢 Column segments: {result['column_segments']}")
        print("="*60)
    else:
        print(f"\n❌ Export failed: {result['message']}")
        sys.exit(1)


if __name__ == "__main__":
    main()